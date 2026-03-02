import logging
from typing import Any

import numpy as np
import openpyxl

from app.parsers.base import BaseParser, ParsedSection

logger = logging.getLogger(__name__)


def _build_formula_map(file_path: str) -> dict[str, Any]:
    """Evaluate all formulas using the formulas library.

    Returns a dict mapping "SheetName!RC" (e.g. "Sheet1!B11") to the computed
    scalar value so the main parser can look up formula results by cell address.
    """
    try:
        import formulas

        xl_model = formulas.ExcelModel().loads(file_path).finish()
        raw = xl_model.calculate()
    except Exception as e:
        logger.warning("formulas library failed, will fall back to formula strings: %s", e)
        return {}

    result: dict[str, Any] = {}
    for key, ranges_obj in raw.items():
        # key looks like "'[filename.xlsx]SheetName'!B11"
        key_str = str(key)
        try:
            # Extract sheet name and cell ref
            sheet_part, cell_ref = key_str.rsplit("!", 1)
            # Remove surrounding quotes and filename bracket
            sheet_part = sheet_part.strip("'")
            bracket_end = sheet_part.find("]")
            if bracket_end != -1:
                sheet_name = sheet_part[bracket_end + 1:]
            else:
                sheet_name = sheet_part

            # Extract scalar value from Ranges object
            val = ranges_obj.value
            if isinstance(val, (list, np.ndarray)):
                # Nested [[value]] structure
                while isinstance(val, (list, np.ndarray)) and len(val) > 0:
                    val = val[0]

            if isinstance(val, (np.integer, np.floating)):
                val = val.item()

            lookup_key = f"{sheet_name}!{cell_ref}"
            result[lookup_key] = val
        except Exception:
            continue

    return result


def _format_value(value: Any) -> str:
    """Format a numeric value for display, removing unnecessary decimals."""
    if isinstance(value, str):
        return value
    if isinstance(value, float):
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
        # Round to 4 decimal places to avoid floating-point noise
        return str(round(value, 4))
    return str(value)


class ExcelParser(BaseParser):
    def parse(self, file_path: str) -> tuple[list[ParsedSection], int]:
        sections: list[ParsedSection] = []

        # Phase 1: compute formulas via the formulas library
        formula_map = _build_formula_map(file_path)
        if formula_map:
            logger.info("Computed %d formula cells", len(formula_map))

        # Phase 2: read workbook for values (data_only=True for cached results)
        wb_data = openpyxl.load_workbook(file_path, data_only=True)

        # Phase 3: read workbook for formula strings (fallback)
        wb_formula = openpyxl.load_workbook(file_path, data_only=False)

        sheet_count = len(wb_data.sheetnames)

        for sheet_name in wb_data.sheetnames:
            sheet_data = wb_data[sheet_name]
            sheet_formula = wb_formula[sheet_name]
            rows: list[str] = []

            for row_data, row_formula in zip(
                sheet_data.iter_rows(), sheet_formula.iter_rows()
            ):
                cells = []
                for cell_data, cell_formula in zip(row_data, row_formula):
                    coord = cell_data.coordinate  # e.g. "B11"

                    if cell_data.value is not None:
                        # Cached computed value exists
                        cells.append(str(cell_data.value).strip())
                    else:
                        # Try formulas library computed value
                        lookup = f"{sheet_name}!{coord}"
                        computed = formula_map.get(lookup)

                        if computed is not None and str(computed).strip() not in ("", "empty"):
                            cells.append(_format_value(computed))
                        elif cell_formula.value is not None:
                            # Last resort: formula string with [수식] prefix
                            cells.append(f"[수식]{cell_formula.value}")
                        else:
                            cells.append("")

                # Strip trailing empty cells (empty columns beyond data range)
                while cells and cells[-1] == "":
                    cells.pop()

                if cells:
                    rows.append(" | ".join(cells))

            if rows:
                sections.append(ParsedSection(
                    text="\n".join(rows),
                    page_number=1,
                    section_title=sheet_name,
                    is_table=True,
                ))

        wb_data.close()
        wb_formula.close()
        return sections, sheet_count
