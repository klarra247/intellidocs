import logging
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from openpyxl import load_workbook

from app.config import settings

logger = logging.getLogger(__name__)

router = APIRouter()

MAX_ROWS = 500
MAX_COLS = 50


@router.get("/preview")
def get_excel_preview(storage_path: str = Query(..., description="서버 내 파일 경로")):
    """Excel 파일을 JSON으로 변환하여 미리보기 데이터를 반환"""

    file_path = Path(storage_path).resolve()

    # 경로 보안: uploads 디렉토리 내 파일만 허용
    allowed_dir = Path(settings.upload_dir).resolve()
    if not str(file_path).startswith(str(allowed_dir)):
        raise HTTPException(status_code=403, detail="허용되지 않은 파일 경로입니다")

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

    try:
        wb = load_workbook(str(file_path), data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"Excel 파일 로드 실패: {file_path}, error={e}")
        raise HTTPException(status_code=500, detail="FILE_CORRUPTED")

    sheets = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            total_rows = 0
            total_cols = 0
            headers = []
            rows = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # 첫 행 = 헤더
                    total_cols = len(row)
                    truncated_cols = total_cols > MAX_COLS
                    col_limit = min(total_cols, MAX_COLS)
                    headers = [str(cell) if cell is not None else "" for cell in row[:col_limit]]
                    total_rows += 1
                else:
                    total_rows += 1
                    if len(rows) < MAX_ROWS:
                        col_limit = min(len(row), MAX_COLS)
                        rows.append(
                            [str(cell) if cell is not None else "" for cell in row[:col_limit]]
                        )

            sheets.append({
                "sheetName": sheet_name,
                "headers": headers,
                "rows": rows,
                "totalRows": total_rows - 1,  # 헤더 제외
                "totalCols": total_cols,
                "truncatedRows": (total_rows - 1) > MAX_ROWS,
                "truncatedCols": total_cols > MAX_COLS,
            })
    finally:
        wb.close()

    return {"sheets": sheets}
